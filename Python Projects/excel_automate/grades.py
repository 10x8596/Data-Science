from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Data
data = {
    "Joe": {
        "math": 65,
        "science": 78,
        "english": 98,
        "gym": 89
    },
    "Bill": {
        "math": 55,
        "science": 72,
        "english": 87,
        "gym": 95
    },
    "Tim": {
        "math": 100,
        "science": 45,
        "english": 75,
        "gym": 92
    },
    "Sally": {
        "math": 30,
        "science": 25,
        "english": 45,
        "gym": 100
    },
    "Jane": {
        "math": 100,
        "science": 100,
        "english": 100,
        "gym": 60
    },
}

# Initialize new workbook
wb = Workbook()

# create first worksheet
ws = wb.active

# Change worksheet title
ws.title = "Grades"

# Write column headings and append data to worksheet
headings = ['Name'] +  list(data["Joe"].keys())
ws.append(headings)

# Loop through data and add it to worksheet
for person in data:
    grades = list(data[person].values())
    ws.append([person] + grades)

# Calculate average of each grades column
for col in range(2, len(data["Joe"]) + 2):
    char = get_column_letter(col)
    ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

# Save the workbook
wb.save('Grades.xlsx')